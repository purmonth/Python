from openpyxl import load_workbook

print("Loading starting")
src_wb = load_workbook('/Users/hung-yuehlin/Python/ExtractData/US1992-1.xlsx')
dest_wb = load_workbook('/Users/hung-yuehlin/Python/ExtractData/US1992-1_done.xlsx')
print("Complete loading")
src_sheet = src_wb.get_sheet_by_name('Sheet2')
dest_sheet = dest_wb.get_sheet_by_name('Sheet1')




wordList = [
"TargetIndustrySector",                             	
"HighTechIndustry",                   	
"TargetState",       
"AcquirorIndustrySector",                           
"AcquirorPrimarySICCode",
"HighTechIndustry", 	
"AcquirorState",     	
"Status",             	 
#"%ofSharesAcq.",	  
"%OwnedAfterTransaction",       	  
"%Sought",
"ValueofTransaction($mil)",
"Premium1DaypriortoannouncementData",
"Premium1WeekpriortoannouncementData",
"Premium4WeekspriortoannouncementData"
]

wordList2 = [
"AcquirorCUSIP",                     
"AcquirorBookValueLTM($Mil)",  	
"AcquirorCashLTM($Mil)",     
"AcqClosing1DayPriortoAnn($)",          	
"AcqClosing1WeekPriortoAnn($)", 	
"AcqClosing4WeeksPriortoAnn($)", 	
"AcqCommonSharesLTM(mil)",    	
"AcqCommonEquityLTM($mil)",    	
"AcqCurrAssetsLTM($Mil)",  	
"AcqCurrLiabilitiesLTM($mil)",	
"AcqEarningsPerShareLTM($)",         	 
"AcqEBITLTM($mil)",             	
"AcqEBITDALTM($mil)",        	
"AcqIndustrySectorCode",	
"AcqNAICCode",       	 
"AcqNetAssets($mil)",          	
"Acquiror'spricepershare",	
"AcquirorState",     	
"AcquirorStateRegionCode",   	
"AcquirorClosingPrice1DayAfterAnnDate($)",	
"AcquirorClosingPrice1WeekAfterAnnDate($)",	
"AcquirorClosingPrice4WeeksAfterAnnDate($)",	
"AcquirorClosingPriceatAnn($)",	
"AcqTickerSymbol",	
"AcqZipCode",       	
"Intrastate",  	
"TargetCUSIP",  	
"TargetIndustrySectorCode",                           	
"TargetNAICCode",	
"TargetZipCode",
]




testList = [
    "DealNumber",
    "TargetImmediateParentPublicStatus",
    "Acq.PublicStatus",
    "Hello"
]

for x in range(len(testList)): 
    print (testList[x])

targetList = []


print("start")


for i in range(1, src_sheet.max_row+1):
    destCol = 1
    print(i/(src_sheet.max_row+1))
    for j in range(1, src_sheet.max_column+1):
        if(src_sheet.cell(row=1, column=j).value in wordList):
            dest_sheet.cell(row=i, column=destCol).value = src_sheet.cell(row=i, column=j).value
            destCol += 1
    

src_wb.save('/Users/hung-yuehlin/Python/ExtractData/US1992-1.xlsx')
dest_wb.save('/Users/hung-yuehlin/Python/ExtractData/US1992-1_done.xlsx')


'''
USxxxx-1 [filtered “date announced” NO blank]
Target Industry Sector                             	
High Tech Industry                   	
Target State       
Acquiror Industry Sector                           
Acquiror Primary SIC Code
High Tech Industry 	
Acquiror State     	
Status             	 
% of Shares Acq.	  
% Owned After Transaction       	  
% Sought	
Value of Transaction ($mil)     	 
Premium 1 Day prior to announcement Data   	 
Premium 1 Week prior to announcement Data   	 
Premium 4 Weeks prior to announcement Data 



'''
